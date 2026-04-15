"""
Automation Coverage Publisher - Azure DevOps Wiki
--------------------------------------------------
Reads AppMigration_TestScenarios_Updated.xlsx and publishes
coverage metrics to an Azure DevOps Wiki page.

Usage:
    python publish_coverage_wiki.py --org https://dev.azure.com/yourorg --project YourProject --pat YOUR_PAT

Or set environment variables:
    set AZURE_DEVOPS_ORG=https://dev.azure.com/yourorg
    set AZURE_DEVOPS_PROJECT=YourProject
    set AZURE_DEVOPS_PAT=your_pat_token
    python publish_coverage_wiki.py
"""
import os
import sys
import json
import argparse
import base64
from datetime import datetime
from collections import defaultdict

import requests
import openpyxl


# ============================================================
# Config
# ============================================================
EXCEL_FILE = os.path.join(os.path.dirname(__file__), "AppMigration_TestScenarios_Updated.xlsx")
WIKI_PAGE_PATH = "/Automation-Coverage"  # Wiki page path

# Sheets to process and their column mappings
SHEETS_CONFIG = {
    "miAccount": {
        "module_col": "Module",
        "description_col": "TestCase Description",
        "auto_col": "Automation / Manual",
        "script_col": "Script",
        "design_col": "Test Design",
        "assigned_col": "AssignedTo",
    },
    "ESS": {
        "module_col": "Module",
        "description_col": "TestCaseDescription",
        "auto_col": "Automation/Manual",
        "script_col": "Script",
        "design_col": "Test Design",
        "assigned_col": "AssignedTo",
    },
    "Clarety": {
        "module_col": "Module",
        "description_col": "Test Case Description",
        "auto_col": "Automation/Manual",
        "script_col": "Script",
        "design_col": "Test Design",
        "assigned_col": "AssignedTo",
    },
}


# ============================================================
# Read Excel Data
# ============================================================
def read_coverage_data(excel_path: str) -> dict:
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    all_data = {}

    for sheet_name, config in SHEETS_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h).strip() if h else "" for h in rows[0]]

        # Find column indices
        def find_col(name):
            for i, h in enumerate(headers):
                if h == name:
                    return i
            return -1

        auto_idx = find_col(config["auto_col"])
        script_idx = find_col(config["script_col"])
        design_idx = find_col(config["design_col"])
        module_idx = find_col(config["module_col"])
        desc_idx = find_col(config["description_col"])
        assigned_idx = find_col(config["assigned_col"])

        test_cases = []
        last_module = ""
        for row in rows[1:]:
            # Skip completely empty rows
            if all(cell is None for cell in row):
                continue

            module = str(row[module_idx]).strip().replace('\xa0', ' ') if module_idx >= 0 and row[module_idx] else last_module
            if module and module != "None":
                last_module = module

            auto_status = str(row[auto_idx]).strip() if auto_idx >= 0 and row[auto_idx] else ""
            if not auto_status or auto_status == "None":
                continue

            script_status = str(row[script_idx]).strip() if script_idx >= 0 and row[script_idx] else ""
            design_status = str(row[design_idx]).strip() if design_idx >= 0 and row[design_idx] else ""
            description = str(row[desc_idx]).strip() if desc_idx >= 0 and row[desc_idx] else ""
            assigned = str(row[assigned_idx]).strip() if assigned_idx >= 0 and row[assigned_idx] else ""

            test_cases.append({
                "module": last_module.strip(),
                "description": description,
                "automation": auto_status,
                "script": script_status,
                "design": design_status,
                "assigned": assigned,
            })

        all_data[sheet_name] = test_cases

    wb.close()
    return all_data


# ============================================================
# Compute Metrics
# ============================================================
def compute_metrics(all_data: dict) -> dict:
    metrics = {}

    for app_name, test_cases in all_data.items():
        total = len(test_cases)
        automated = sum(1 for tc in test_cases if tc["automation"].lower() == "automation")
        manual = sum(1 for tc in test_cases if tc["automation"].lower() == "manual")
        script_done = sum(1 for tc in test_cases if tc["script"].lower() == "done")
        design_done = sum(1 for tc in test_cases if tc["design"].lower() == "done")

        # Module breakdown
        module_stats = defaultdict(lambda: {"total": 0, "automated": 0, "manual": 0, "script_done": 0})
        for tc in test_cases:
            m = tc["module"] or "Uncategorized"
            module_stats[m]["total"] += 1
            if tc["automation"].lower() == "automation":
                module_stats[m]["automated"] += 1
            if tc["automation"].lower() == "manual":
                module_stats[m]["manual"] += 1
            if tc["script"].lower() == "done":
                module_stats[m]["script_done"] += 1

        metrics[app_name] = {
            "total": total,
            "automated": automated,
            "manual": manual,
            "automation_pct": round(automated / total * 100, 1) if total > 0 else 0,
            "script_done": script_done,
            "script_pct": round(script_done / total * 100, 1) if total > 0 else 0,
            "design_done": design_done,
            "design_pct": round(design_done / total * 100, 1) if total > 0 else 0,
            "modules": dict(module_stats),
        }

    # Grand totals
    grand_total = sum(m["total"] for m in metrics.values())
    grand_automated = sum(m["automated"] for m in metrics.values())
    grand_manual = sum(m["manual"] for m in metrics.values())
    grand_script = sum(m["script_done"] for m in metrics.values())
    grand_design = sum(m["design_done"] for m in metrics.values())

    metrics["_overall"] = {
        "total": grand_total,
        "automated": grand_automated,
        "manual": grand_manual,
        "automation_pct": round(grand_automated / grand_total * 100, 1) if grand_total > 0 else 0,
        "script_done": grand_script,
        "script_pct": round(grand_script / grand_total * 100, 1) if grand_total > 0 else 0,
        "design_done": grand_design,
        "design_pct": round(grand_design / grand_total * 100, 1) if grand_total > 0 else 0,
    }

    return metrics


# ============================================================
# Generate Markdown
# ============================================================
def generate_markdown(metrics: dict) -> str:
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    overall = metrics["_overall"]

    lines = []
    lines.append("# Automation Coverage Report")
    lines.append(f"_Last updated: {now}_\n")

    # Overall summary
    lines.append("## Overall Summary\n")
    lines.append("| Metric | Count | Percentage |")
    lines.append("|--------|------:|------------|")
    lines.append(f"| **Total Test Cases** | **{overall['total']}** | - |")
    lines.append(f"| Automated | {overall['automated']} | {overall['automation_pct']}% |")
    lines.append(f"| Manual | {overall['manual']} | {round(100 - overall['automation_pct'], 1)}% |")
    lines.append(f"| Scripts Completed | {overall['script_done']} | {overall['script_pct']}% |")
    lines.append(f"| Test Design Completed | {overall['design_done']} | {overall['design_pct']}% |")
    lines.append("")

    # Progress bar (text-based)
    def progress_bar(pct):
        filled = int(pct / 5)
        empty = 20 - filled
        return f"`[{'#' * filled}{'.' * empty}]` {pct}%"

    lines.append("### Automation Progress\n")
    lines.append(f"- Overall: {progress_bar(overall['automation_pct'])}")
    lines.append(f"- Scripts Ready: {progress_bar(overall['script_pct'])}")
    lines.append(f"- Test Design Ready: {progress_bar(overall['design_pct'])}")
    lines.append("")

    # Per-application breakdown
    lines.append("## Application Breakdown\n")
    lines.append("| Application | Total | Automated | Manual | Coverage % | Scripts Done |")
    lines.append("|-------------|------:|----------:|-------:|-----------:|-------------:|")

    for app_name in ["miAccount", "ESS", "Clarety"]:
        if app_name not in metrics:
            continue
        m = metrics[app_name]
        lines.append(f"| **{app_name}** | {m['total']} | {m['automated']} | {m['manual']} | {m['automation_pct']}% | {m['script_done']} |")

    lines.append("")

    # Module-level detail per app
    for app_name in ["miAccount", "ESS", "Clarety"]:
        if app_name not in metrics:
            continue
        m = metrics[app_name]
        lines.append(f"## {app_name} - Module Breakdown\n")
        lines.append("| Module | Total | Automated | Manual | Scripts Done |")
        lines.append("|--------|------:|----------:|-------:|-------------:|")

        for mod_name, mod_stats in m["modules"].items():
            display_name = mod_name.strip() if mod_name else "Other"
            lines.append(f"| {display_name} | {mod_stats['total']} | {mod_stats['automated']} | {mod_stats['manual']} | {mod_stats['script_done']} |")

        lines.append("")

    lines.append("---")
    lines.append(f"_Generated automatically from AppMigration_TestScenarios_Updated.xlsx_")

    return "\n".join(lines)


# ============================================================
# Publish to Azure DevOps Wiki
# ============================================================
def publish_to_wiki(org_url: str, project: str, pat: str, wiki_page_path: str, content: str):
    """Publish markdown content to Azure DevOps Wiki page."""
    # First, try to get the project wiki
    auth = base64.b64encode(f":{pat}".encode()).decode()
    headers = {
        "Authorization": f"Basic {auth}",
        "Content-Type": "application/json",
    }

    # List wikis to find the project wiki
    wiki_url = f"{org_url}/{project}/_apis/wiki/wikis?api-version=7.1"
    resp = requests.get(wiki_url, headers=headers)

    if resp.status_code != 200:
        print(f"Failed to list wikis: {resp.status_code} {resp.text}")
        # Try to create a project wiki
        print("Attempting to create project wiki...")
        create_resp = requests.post(
            f"{org_url}/{project}/_apis/wiki/wikis?api-version=7.1",
            headers=headers,
            json={
                "name": f"{project}.wiki",
                "type": "projectWiki",
            }
        )
        if create_resp.status_code in (200, 201):
            wiki_id = create_resp.json()["id"]
            print(f"Created project wiki: {wiki_id}")
        else:
            print(f"Failed to create wiki: {create_resp.status_code} {create_resp.text}")
            return False
    else:
        wikis = resp.json().get("value", [])
        if not wikis:
            print("No wikis found. Creating project wiki...")
            create_resp = requests.post(
                f"{org_url}/{project}/_apis/wiki/wikis?api-version=7.1",
                headers=headers,
                json={
                    "name": f"{project}.wiki",
                    "type": "projectWiki",
                }
            )
            if create_resp.status_code in (200, 201):
                wiki_id = create_resp.json()["id"]
            else:
                print(f"Failed to create wiki: {create_resp.status_code} {create_resp.text}")
                return False
        else:
            wiki_id = wikis[0]["id"]
            print(f"Using wiki: {wikis[0]['name']} (id: {wiki_id})")

    # Create or update the wiki page
    page_url = f"{org_url}/{project}/_apis/wiki/wikis/{wiki_id}/pages?path={wiki_page_path}&api-version=7.1"

    # Try to get existing page first (to get ETag for update)
    get_resp = requests.get(page_url, headers=headers)

    if get_resp.status_code == 200:
        # Page exists - update it
        etag = get_resp.headers.get("ETag", "")
        update_headers = {**headers, "If-Match": etag}
        put_resp = requests.put(
            page_url,
            headers=update_headers,
            json={"content": content}
        )
        if put_resp.status_code == 200:
            print(f"Wiki page updated: {wiki_page_path}")
            return True
        else:
            print(f"Failed to update page: {put_resp.status_code} {put_resp.text}")
            return False
    else:
        # Page doesn't exist - create it
        put_resp = requests.put(
            page_url,
            headers=headers,
            json={"content": content}
        )
        if put_resp.status_code in (200, 201):
            print(f"Wiki page created: {wiki_page_path}")
            return True
        else:
            print(f"Failed to create page: {put_resp.status_code} {put_resp.text}")
            return False


# ============================================================
# Main
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="Publish automation coverage to Azure DevOps Wiki")
    parser.add_argument("--org", default=os.environ.get("AZURE_DEVOPS_ORG", ""), help="Azure DevOps org URL")
    parser.add_argument("--project", default=os.environ.get("AZURE_DEVOPS_PROJECT", ""), help="Project name")
    parser.add_argument("--pat", default=os.environ.get("AZURE_DEVOPS_PAT", ""), help="Personal Access Token")
    parser.add_argument("--excel", default=EXCEL_FILE, help="Path to Excel file")
    parser.add_argument("--preview", action="store_true", help="Preview markdown without publishing")
    args = parser.parse_args()

    # Read and compute
    print(f"Reading coverage data from: {args.excel}")
    all_data = read_coverage_data(args.excel)

    for app, cases in all_data.items():
        print(f"  {app}: {len(cases)} test cases")

    metrics = compute_metrics(all_data)
    markdown = generate_markdown(metrics)

    if args.preview:
        print("\n" + "=" * 60)
        print("PREVIEW - Wiki Page Content")
        print("=" * 60)
        print(markdown)
        return

    # Validate Azure DevOps params
    if not args.org or not args.project or not args.pat:
        print("\nMissing Azure DevOps configuration. Use --preview to see output without publishing.")
        print("Required: --org, --project, --pat (or set AZURE_DEVOPS_ORG, AZURE_DEVOPS_PROJECT, AZURE_DEVOPS_PAT)")
        # Still save markdown locally
        local_path = os.path.join(os.path.dirname(args.excel), "automation_coverage.md")
        with open(local_path, "w") as f:
            f.write(markdown)
        print(f"\nMarkdown saved locally: {local_path}")
        return

    # Publish
    success = publish_to_wiki(args.org, args.project, args.pat, WIKI_PAGE_PATH, markdown)
    if success:
        print(f"\nDone! View at: {args.org}/{args.project}/_wiki/wikis/{args.project}.wiki{WIKI_PAGE_PATH}")

    # Also save locally as backup
    local_path = os.path.join(os.path.dirname(args.excel), "automation_coverage.md")
    with open(local_path, "w") as f:
        f.write(markdown)
    print(f"Markdown also saved locally: {local_path}")


if __name__ == "__main__":
    main()
