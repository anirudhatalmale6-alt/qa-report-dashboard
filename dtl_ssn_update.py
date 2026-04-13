"""
DTL SSN Update Script
---------------------
Searches each SSN from the tracker Excel, checks if BadFormat or Suspended,
updates the required fields, and writes status back to the Excel sheet.

Usage (in TestPlan):
  - SheetName: Report_0708  (or Report_0715, Report_0729)
  - DataReference: ALL  (or range like 1-100, or single like 5)

Or run standalone:
  pytest dtl_ssn_update.py --env uat75 --mode prod --report Report_0708
"""
import os
import re
import time
import logging
from datetime import datetime
from openpyxl import load_workbook
from playwright.sync_api import Page, expect

logger = logging.getLogger(__name__)

# ============================================================
# CONFIG - Update these paths to match your local setup
# ============================================================
TRACKER_PATH = os.path.join(os.path.dirname(__file__), "DTL_SSN_Tracker.xlsx")


# ============================================================
# Excel Helper: Read SSNs from a report sheet
# ============================================================
def load_ssns_from_sheet(sheet_name: str, tracker_path: str = TRACKER_PATH) -> list[dict]:
    """
    Load all SSN rows from the given sheet.
    Returns list of dicts with keys: row_num, SSN, RecordType, etc.
    Skips rows where ActionTaken is already filled (resume support).
    """
    wb = load_workbook(tracker_path)
    ws = wb[sheet_name]

    headers = [cell.value for cell in ws[1]]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = dict(zip(headers, row))
        record["_row_num"] = row_idx

        # Skip already-processed rows (resume support)
        if record.get("ActionTaken") and str(record["ActionTaken"]).strip():
            logger.info(f"Row {row_idx} SSN {record.get('SSN')} already processed, skipping")
            continue

        if not record.get("SSN"):
            continue

        rows.append(record)

    wb.close()
    logger.info(f"Loaded {len(rows)} unprocessed SSNs from sheet '{sheet_name}'")
    return rows


def update_tracker_row(sheet_name: str, row_num: int, updates: dict, tracker_path: str = TRACKER_PATH):
    """
    Write status updates back to the tracker Excel.
    updates: dict of column_name -> value, e.g. {"ActualStatus": "Suspended", "ActionTaken": "Updated"}
    """
    wb = load_workbook(tracker_path)
    ws = wb[sheet_name]

    headers = {cell.value: cell.column for cell in ws[1]}

    for col_name, value in updates.items():
        if col_name in headers:
            ws.cell(row=row_num, column=headers[col_name], value=value)

    wb.save(tracker_path)
    wb.close()


# ============================================================
# Page Actions - CUSTOMIZE THESE for your application
# ============================================================

def search_ssn(page: Page, ssn: str) -> bool:
    """
    Search for an SSN in the application.
    Returns True if found, False if not found.

    TODO: Update selectors to match your application's search page.
    """
    # Example - update these XPaths to match your app:
    # page.locator("//input[@name='ssn']").fill(ssn)
    # page.locator("//input[@type='submit' and @value='Search']").click()
    # page.wait_for_load_state("networkidle")

    # Check if SSN was found (no "not found" message)
    # return not page.locator("//div[contains(@class,'error')]").is_visible()

    raise NotImplementedError("Update search_ssn() with your application's search selectors")


def get_record_status(page: Page) -> dict:
    """
    After searching an SSN, read its current status from the page.
    Returns dict with: record_type, status (BadFormat/Suspended/Active), etc.

    TODO: Update selectors to match your application's result page.
    """
    # Example:
    # record_type = page.locator("//span[@id='recordType']").text_content().strip()
    # status = page.locator("//span[@id='status']").text_content().strip()
    # return {"record_type": record_type, "status": status}

    raise NotImplementedError("Update get_record_status() with your application's selectors")


def update_record_fields(page: Page, record_type: str, ssn_data: dict) -> str:
    """
    Perform the field updates based on record type (DTL2 or DTL4).
    Returns action description string.

    TODO: Implement the actual field update logic.

    For DTL2:
      - Check Record Type field
      - If BadFormat: update XYZ fields
      - If Suspended: update ABC fields

    For DTL4:
      - Similar but different field set
    """
    # Example flow:
    # if record_type == "DTL2":
    #     page.locator("//select[@name='field1']").select_option("value1")
    #     page.locator("//input[@name='field2']").fill("new_value")
    #     page.locator("//input[@type='submit']").click()
    #     page.wait_for_load_state("networkidle")
    #     return "Updated DTL2 fields"
    # elif record_type == "DTL4":
    #     ...
    #     return "Updated DTL4 fields"

    raise NotImplementedError("Update update_record_fields() with your field update logic")


# ============================================================
# Main Processing Loop
# ============================================================
def process_report(page: Page, sheet_name: str, tracker_path: str = TRACKER_PATH):
    """
    Main loop: iterate all SSNs in the sheet, search, check, update, track.
    Safe to re-run - skips already-processed rows.
    """
    ssn_rows = load_ssns_from_sheet(sheet_name, tracker_path)
    total = len(ssn_rows)
    logger.info(f"Processing {total} SSNs for {sheet_name}")

    for idx, ssn_data in enumerate(ssn_rows, 1):
        ssn = str(ssn_data["SSN"]).strip()
        row_num = ssn_data["_row_num"]
        expected_type = str(ssn_data.get("RecordType", "")).strip()
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        logger.info(f"[{idx}/{total}] Processing SSN: {ssn} (Row {row_num})")

        try:
            # Step 1: Search for the SSN
            found = search_ssn(page, ssn)
            if not found:
                update_tracker_row(sheet_name, row_num, {
                    "ActualStatus": "NOT FOUND",
                    "ActionTaken": "Skipped",
                    "Timestamp": timestamp,
                    "Notes": "SSN not found in system"
                }, tracker_path)
                logger.warning(f"SSN {ssn} not found, skipping")
                continue

            # Step 2: Read the record status
            status_info = get_record_status(page)
            actual_status = status_info.get("status", "Unknown")
            record_type = status_info.get("record_type", expected_type)

            # Step 3: Only process DTL2 and DTL4
            if record_type not in ("DTL2", "DTL4"):
                update_tracker_row(sheet_name, row_num, {
                    "ActualStatus": actual_status,
                    "ActionTaken": "Skipped",
                    "Timestamp": timestamp,
                    "Notes": f"Record type is {record_type}, not DTL2/DTL4"
                }, tracker_path)
                logger.info(f"SSN {ssn} is {record_type}, skipping (not DTL2/DTL4)")
                continue

            # Step 4: Perform the update
            action = update_record_fields(page, record_type, ssn_data)

            # Step 5: Write success back to tracker
            update_tracker_row(sheet_name, row_num, {
                "RecordType": record_type,
                "ActualStatus": actual_status,
                "ActionTaken": action,
                "Timestamp": timestamp,
                "Notes": "OK"
            }, tracker_path)
            logger.info(f"SSN {ssn} -> {actual_status} -> {action}")

        except Exception as e:
            # Write error back to tracker so we can see what failed
            update_tracker_row(sheet_name, row_num, {
                "ActionTaken": "Error",
                "Timestamp": timestamp,
                "Notes": str(e)[:200]
            }, tracker_path)
            logger.error(f"SSN {ssn} failed: {e}")
            continue

    logger.info(f"Finished processing {sheet_name}: {total} SSNs")


# ============================================================
# Pytest Integration (optional - can also call process_report directly)
# ============================================================
def test_dtl_ssn_update(test_context):
    """
    Pytest entry point. Reads sheet name from test_context.
    Configure in TestPlan:
      TestName: DTL_SSN_Update
      TestMethod: test_dtl_ssn_update
      SheetName: Report_0708
      DataReference: ALL
    """
    page = test_context["page"]
    sheet_name = test_context.get("sheet_name", "Report_0708")

    process_report(page, sheet_name)
