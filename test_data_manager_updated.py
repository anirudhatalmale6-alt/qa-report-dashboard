# lib/test_data_manager.py
from datetime import datetime
import csv
import os
import time
from openpyxl import load_workbook
from lib.logger import logger
from lib.constants import TEST_PLAN_PATH
from filelock import FileLock
from lib.constants import resolve_column

LOCK_FILE = TEST_PLAN_PATH + ".lock"

# ---------------------------------------------------------
# Results CSV Configuration
# ---------------------------------------------------------
# Results directory (created at session start)
RESULTS_DIR = "test-results"
# Lock file for CSV writes (thread-safe for xdist parallel)
CSV_LOCK_FILE = os.path.join(RESULTS_DIR, "results.lock")


def _get_results_csv_path(sheet_name: str) -> str:
    """Get the results CSV file path for a given sheet."""
    return os.path.join(RESULTS_DIR, f"results_{sheet_name}.csv")


def _ensure_results_dir():
    """Create results directory if it doesn't exist."""
    os.makedirs(RESULTS_DIR, exist_ok=True)


def init_results_csv(sheet_name: str, file_path: str = TEST_PLAN_PATH):
    """
    Initialize a results CSV for a sheet by copying DataReference and MemberID
    from the source Excel. Called once at session start.
    """
    _ensure_results_dir()
    csv_path = _get_results_csv_path(sheet_name)

    # If CSV already exists (e.g., from a previous partial run), skip
    if os.path.exists(csv_path):
        return

    # Read headers and DataReference values from Excel
    lock = FileLock(LOCK_FILE, timeout=60)
    with lock:
        wb = load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return
        sheet = wb[sheet_name]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        headers = [h for h in headers if h is not None]

        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            rows.append(row_data)
        wb.close()

    # Write CSV with all headers and data
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row_data in rows:
            writer.writerow(row_data)

    logger.info(f"Initialized results CSV: {csv_path} ({len(rows)} rows, {len(headers)} columns)")


# ---------------------------------------------------------
# File writable check (unchanged)
# ---------------------------------------------------------
def check_file_writable(file_path: str = TEST_PLAN_PATH):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found at {file_path}")
    try:
        with open(file_path, "a+"):
            pass
    except PermissionError:
        raise PermissionError(
            f"File '{file_path}' is locked or opened in another program. "
            f"Please close it and re-run tests."
        )


# ---------------------------------------------------------
# Excel read-only functions (unchanged)
# ---------------------------------------------------------
def safe_update_excel(update_func, file_path: str = TEST_PLAN_PATH):
    """Keep for backward compatibility but no longer used for writes."""
    lock = FileLock(LOCK_FILE, timeout=600)
    with lock:
        wb = load_workbook(file_path)
        update_func(wb)
        wb.save(file_path)


def load_test_data(sheet_name: str, data_ref: str, file_path: str) -> dict:
    """Read test data from Excel (read-only, no corruption risk)."""
    lock = FileLock(LOCK_FILE, timeout=600)
    with lock:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found at {file_path}")
        wb = load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in Excel file")
        sheet = wb[sheet_name]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        col_map = {h: idx for idx, h in enumerate(headers)}
        if "DataReference" not in col_map:
            raise KeyError("'DataReference' column is missing in sheet")
        for row in sheet.iter_rows(min_row=2, values_only=False):
            if str(row[col_map["DataReference"]].value).strip() == str(data_ref).strip():
                return {headers[i]: row[i].value for i in range(len(headers))}
        raise ValueError(f"No data found for DataReference: '{data_ref}' in sheet '{sheet_name}'")


# ---------------------------------------------------------
# CSV-based write functions (replaces Excel writes)
# ---------------------------------------------------------
def write_to_cell(row_ref_value: str, col_header: str, value,
                  sheet_name: str, file_path: str = TEST_PLAN_PATH,
                  ref_col: str = 'DataReference'):
    """
    Write a value to the results CSV instead of Excel.
    Same signature as before - no changes needed in calling code.
    """
    _ensure_results_dir()
    csv_path = _get_results_csv_path(sheet_name)
    csv_lock = FileLock(CSV_LOCK_FILE, timeout=600)

    # Resolve column name (old -> new mapping)
    col_header = resolve_column(col_header)

    with csv_lock:
        # If CSV doesn't exist yet, initialize it from Excel
        if not os.path.exists(csv_path):
            init_results_csv(sheet_name, file_path)

        # If still doesn't exist (sheet not found), create minimal CSV
        if not os.path.exists(csv_path):
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=[ref_col, col_header, 'Timestamp'])
                writer.writeheader()

        # Read existing CSV
        rows = []
        headers = []
        with open(csv_path, 'r', newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            headers = list(reader.fieldnames) if reader.fieldnames else []
            rows = list(reader)

        # Add column if it doesn't exist
        if col_header not in headers:
            headers.append(col_header)
            logger.info(f"Added new column '{col_header}' to results CSV for sheet '{sheet_name}'")

        # Ensure Timestamp column exists
        if 'Timestamp' not in headers:
            headers.append('Timestamp')

        # Find and update the matching row
        found = False
        for row in rows:
            if str(row.get(ref_col, '')).strip() == str(row_ref_value).strip():
                row[col_header] = value
                row['Timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                found = True
                logger.info(f"CSV updated: {ref_col}={row_ref_value}, {col_header}={value}")
                break

        if not found:
            # Create new row if DataReference not found (edge case)
            new_row = {h: '' for h in headers}
            new_row[ref_col] = row_ref_value
            new_row[col_header] = value
            new_row['Timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            rows.append(new_row)
            logger.warning(f"CSV: Created new row for {ref_col}={row_ref_value} in sheet '{sheet_name}'")

        # Write back the full CSV
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(rows)


def update_test_status(test_name: str, status: str,
                       file_path: str = TEST_PLAN_PATH, sheet_name: str = 'TestPlan'):
    """
    Update test status in results CSV instead of Excel.
    """
    _ensure_results_dir()
    csv_path = _get_results_csv_path(sheet_name)
    csv_lock = FileLock(CSV_LOCK_FILE, timeout=600)

    with csv_lock:
        # Initialize CSV from Excel if needed
        if not os.path.exists(csv_path):
            init_results_csv(sheet_name, file_path)

        if not os.path.exists(csv_path):
            # TestPlan sheet not found - create minimal CSV
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=['TestName', 'Status', 'TimeStamp'])
                writer.writeheader()

        # Read existing CSV
        rows = []
        headers = []
        with open(csv_path, 'r', newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            headers = list(reader.fieldnames) if reader.fieldnames else []
            rows = list(reader)

        # Ensure required columns
        for col in ['Status', 'TimeStamp']:
            if col not in headers:
                headers.append(col)

        # Find and update
        found = False
        for row in rows:
            if str(row.get('TestName', '')).strip() == str(test_name).strip():
                row['Status'] = status
                row['TimeStamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                found = True
                logger.info(f"Test status updated in CSV: {test_name} => {status}")
                break

        if not found:
            new_row = {h: '' for h in headers}
            new_row['TestName'] = test_name
            new_row['Status'] = status
            new_row['TimeStamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            rows.append(new_row)
            logger.warning(f"TestName '{test_name}' not found in CSV, created new row")

        # Write back
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(rows)


def wait_for_test_status(test_name: str, status: str, timeout: int = 900,
                         file_path: str = TEST_PLAN_PATH, sheet_name: str = 'TestPlan',
                         poll_interval: int = 10):
    """
    Wait for a test to reach a specific status.
    Now reads from results CSV instead of Excel.
    """
    csv_path = _get_results_csv_path(sheet_name)
    csv_lock = FileLock(CSV_LOCK_FILE, timeout=600)
    start_time = time.time()
    logger.info(f"Waiting for {test_name} to reach status '{status}'")

    while True:
        with csv_lock:
            if os.path.exists(csv_path):
                with open(csv_path, 'r', newline='', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        if str(row.get('TestName', '')).strip() == str(test_name).strip():
                            current_status = str(row.get('Status', '')).strip()
                            if current_status.lower() == status.lower():
                                logger.info(f"{test_name} reached status '{status}'")
                                return True
                            break

        if time.time() - start_time > timeout:
            raise TimeoutError(
                f"Timeout: '{test_name}' did not reach status '{status}' in {timeout} seconds."
            )
        time.sleep(poll_interval)
